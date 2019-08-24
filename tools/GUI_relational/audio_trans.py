#-*-encoding:utf-8-*-

import os
import sys
import tkinter as tk
from  tkinter import filedialog
from tkinter import ttk
import time
import numpy as np
import openpyxl
import subprocess

# # 创建全局变量
# file_num = 0
# file_cnt = 0
'''
    开始转换音频格式任务响应
'''
def begin_trans():
    # global file_cnt
    # global file_num
    src_format = src_audio_format.get()
    tgt_format = tgt_audio_format.get()
    print(src_file_path.get())
    print(tgt_file_path.get())
    if os.path.isdir(src_file_path.get()):
        for roots, folder, files in os.walk(src_file_path.get()):
            file_cnt = 0
            for file in files:
                # 统计文件个数
                file_num = len(files)
                # 记录文件index
                file_cnt += 1
                audio_path = os.path.join(roots,file)
                save_path = os.path.join(tgt_file_path.get(),file.split('.')[0] + '.' + tgt_format)
                audio_trans(src_format, tgt_format, audio_path, save_path)
                trans_progress(canvas, file_num, file_cnt)
    else:
        save_path = os.path.join(tgt_file_path.get(), os.path.split(src_file_path.get())[-1].split('.')[0] + '.' + tgt_format)
        audio_trans(src_format, tgt_format, src_file_path.get(), save_path)

'''
    暂停转换音频格式任务响应
'''
def pause_trans():
    print("2222")
    os.system("pause")

'''
    取消转换音频格式任务响应
'''
def cancel_trans():
    print("3333")
    os._exit(1)

'''
    显示转换进度条
'''
def trans_progress(canvas, file_num, file_cnt):
    # 填充进度条
    fill_line = canvas.create_rectangle(1.5, 1.5, 0, 23, width=0, fill="green")
    x = file_num  # 未知变量，可更改
    n = 430 * file_cnt / x   # 465是矩形填充满的次数
    canvas.coords(fill_line, (0, 0, n, 60))
    window.update()
    time.sleep(0.02)  # 控制进度条流动的速度
    # for i in range(x):
    #     n = n + 465 / x
    #     canvas.coords(fill_line, (0, 0, n, 60))
    #     window.update()
    #     time.sleep(0.02)  # 控制进度条流动的速度

'''
    创建窗口
'''
def create_window(title,win_width,win_height):
    # 第1步，实例化object，建立窗口window
    window = tk.Tk()
    # 第2步，给窗口的可视化起名字
    window.title(title)
    # 第3步，设定窗口的大小(长 * 宽)
    window.geometry("%sx%s" %(win_width,win_height))  # 这里的乘是小x
    return window

'''
    创建frame
'''
def set_frame(window):
    frame = tk.Frame(window)
    #放置
    frame.pack()
    return frame


'''
    选择单个文件
'''
def select_src_file():
    # 选择文件path_接收文件地址
    src_file_path = tk.filedialog.askopenfilename()
    if src_file_path != '':
        print(src_file_path)
        entry_src_file_path_input.delete(0, tk.END)  # 清空文本框
        entry_src_file_path_input.insert(tk.END, src_file_path)  # 填充文本框
'''
    选择文件夹
'''
def select_src_folder():
    src_file_path = tk.filedialog.askdirectory()
    if src_file_path != '':
        print(src_file_path)
        entry_src_file_path_input.delete(0,tk.END)          #清空文本框
        entry_src_file_path_input.insert(tk.END,src_file_path)  #填充文本框

'''
    选择目标文件夹
'''
def select_tgt_folder():
    tgt_file_path = tk.filedialog.askdirectory()
    if tgt_file_path != '':
        print(tgt_file_path)
        entry_tgt_file_path_input.delete(0,tk.END)          #清空文本框
        entry_tgt_file_path_input.insert(tk.END,tgt_file_path)  #填充文本框


'''
    音频类型转化，使用ffmpeg工具
'''
def audio_trans(src_format,tgt_format,src_file_path,tgt_file_path):
    if tgt_format == 'wav':
        try:
            cmd_trans = "ffmpeg -f s16be -ar 16000 -ac 1 -acodec pcm_s16be -i %s %s" %(src_file_path, tgt_file_path)
            print(cmd_trans)
            # os.system(cmd_trans)
            res = subprocess.call(cmd_trans, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                                  stderr=subprocess.PIPE)
            return True
        except:
            return False
    else:
        if src_format == 'wav' and tgt_format == 'pcm':
            return  wav_to_pcm(src_file_path, tgt_file_path)
    # wav转pcm
    cmd_trans = r"ffmpeg -i %s -f s16be -ar 8000 -acodec pcm_s16be %s"
    #pcm转wav
    cmd_trans = r"ffmpeg -f s16be -ar 8000 -ac 2 -acodec pcm_s16be -i %s output.wav"

'''
    wav 转 pcm，去除44字节头信息
'''
def wav_to_pcm(src_file,tgt_file):
    try:
        f = open(src_file, 'rb')
        f.seek(0)
        f.read(44)
        data = np.fromfile(f, dtype=np.int16)
        data.tofile(tgt_file)
        return True
    except:
        return False

###################################### Audio split ##############################################################
'''
    选择单个文件
'''
def select_src_file_split():
    # 选择文件path_接收文件地址
    src_file_path_split_split = tk.filedialog.askopenfilename()
    if src_file_path_split_split != '':
        print(src_file_path_split_split)
        entry_src_file_path_input_split.delete(0, tk.END)  # 清空文本框
        entry_src_file_path_input_split.insert(tk.END, src_file_path_split_split)  # 填充文本框
'''
    选择文件夹
'''
def select_src_folder_split():
    src_file_path_split_split = tk.filedialog.askdirectory()
    if src_file_path_split_split != '':
        print(src_file_path_split_split)
        entry_src_file_path_input_split.delete(0,tk.END)          #清空文本框
        entry_src_file_path_input_split.insert(tk.END,src_file_path_split_split)  #填充文本框

'''
    选择目标文件夹
'''
def select_tgt_folder_split():
    tgt_file_path_split = tk.filedialog.askdirectory()
    if tgt_file_path_split != '':
        print(tgt_file_path_split)
        entry_tgt_file_path_input_split.delete(0,tk.END)          #清空文本框
        entry_tgt_file_path_input_split.insert(tk.END,tgt_file_path_split)  #填充文本框

'''
    选择时间分片excel文件
'''
def select_segment_file_split():
    segment_file_split_path = tk.filedialog.askopenfilename(filetypes=[("文本文件", "*.xlsx"),])
    if segment_file_split_path != '':
        print(segment_file_split_path)
        entry_segment_file_path.delete(0,tk.END)          #清空文本框
        entry_segment_file_path.insert(tk.END,segment_file_split_path)  #填充文本框

'''
    开始音频分割响应
'''
def begin_split():
    if segment_file_path.get() != '' and first_audio_name.get()!= '':   # 批量音频文件切割,按照segment文件进行切割
        file_dict = get_video_info(segment_file_path.get(), first_audio_name.get())
        cut_audio(file_dict, src_file_path_split.get(), tgt_file_path_split.get())
    elif os.path.isdir(src_file_path_split.get()): #批量音频文件切割，按照begin end time进行切割
        # 需要对begin和end时间进行判断,判断格式是否为00:00:00 ，begin是否小于end
        if len(begin_time.get().split(':')) < 3:
            begin = '00:' * (3 - len(begin_time.get().split(':'))) + begin_time.get()
        if len(end_time.get().split(':')) < 3:
            end = '00:' * (3 - len(end_time.get().split(':'))) + end_time.get()
        if time_trans(begin) >= time_trans(end):
            # 退出当前转换，并打印相关信息
            print('The time of begin and end is error, please check it.')
        for roots, folders, files in os.walk(src_file_path_split.get()):
            file_num = len(files)
            file_cnt = 0
            for file in files:
                audio_path = os.path.join(roots, file)
                save_path = os.path.join(tgt_file_path_split.get(), file)
                ffmpeg_cmd = "ffmpeg -i %s -ss %s -to %s %s" % (audio_path, begin, end, save_path)
                res = subprocess.call(ffmpeg_cmd, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                                      stderr=subprocess.PIPE)
                file_cnt += 1
                if res != 0:
                    print('%s is failed.' % file)
                trans_progress(canvas_split, file_num, file_cnt)

    else:   #单音频文件切割
        # 需要对begin和end时间进行判断,判断格式是否为00:00:00 ，begin是否小于end
        if len(begin_time.get().split(':')) < 3:
            begin = '00:' * (3 - len(begin_time.get().split(':'))) + begin_time.get()
        if len( end_time.get().split(':')) < 3:
            end = '00:' * (3 - len( end_time.get().split(':'))) +  end_time.get()
        if time_trans(begin) >= time_trans(end):
            # 退出当前转换，并打印相关信息
            print('The time of begin and end is error, please check it.')
        ffmpeg_cmd = "ffmpeg -i %s -ss %s -to %s %s" % (src_file_path_split.get(),
                                                                     begin, end,
                                                                     os.path.join(tgt_file_path_split.get(), os.path.split(src_file_path_split.get())[-1]))
        # os.system(ffmpeg_cmd)
        res = subprocess.call(ffmpeg_cmd, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE)



'''
    获取有效音频列表
    获取音频对应时间片段，保存格式如下
    {
        video_name:{
            'segment_1':{
                begin:begin,
                end:end
            }
        }
    }
'''
def get_video_info(excel_file,first_name):
    video_list = [] #有效音频列表
    wb = openpyxl.load_workbook(excel_file)  # 读取xlsx文件
    sheet = wb['Sheet1']        #读取工作表
    rows = sheet.max_row  # 行数
    print(rows)
    # columns = sheet.max_column                      #列数
    file_name = first_name
    segment_cnt = 0
    file_dict = {}
    segment_dict = {}
    for i in range(2, rows + 1):
        video_name = sheet.cell(row=i, column=1).value  #音频名
        begin = sheet.cell(row=i, column=2).value       #segment begin
        end = sheet.cell(row=i, column=3).value        #segment end
        # 需要对begin和end时间进行判断,判断格式是否为00:00:00 ，begin是否小于end
        if len(begin.split(':')) < 3:
            begin = '00:' * (3 - len(begin.split(':'))) + begin
        if len(end.split(':')) < 3:
            end = '00:' * (3 - len(end.split(':'))) + end
        if time_trans(begin) >= time_trans(end):
            # 退出当前转换，并打印相关信息
            print('The time of begin and end is error, please check it.')
        if video_name is not None and video_name == first_name:
            # segment_dict.setdefault(str(segment_cnt), {'begin': begin, 'end': end})
            video_list.append(video_name)
            pass
        elif video_name is not None:
            video_list.append(video_name)
            #保存到字典
            file_dict.setdefault(file_name,segment_dict)
            file_name = video_name
            segment_cnt = 0
            segment_dict = {}
            # segment_dict.setdefault(str(segment_cnt), {'begin': begin, 'end': end})
        else:
            pass
        segment_dict.setdefault(str(segment_cnt),{'begin':begin,'end':end})
        segment_cnt += 1
        # video_list.append(video_name)
    return file_dict

'''
    调用ffmpeg工具，切分音频
'''
def cut_audio(file_dict,src_video,tgt_video):
    file_num = len(file_dict.keys())
    file_cnt = 0
    for key in file_dict.keys():
        cnt = 0
        for seg in file_dict.get(key).keys():
            begin = file_dict.get(key).get(seg)['begin']
            end = file_dict.get(key).get(seg)['end']
            ffmpeg_cmd = "ffmpeg -i %s -ss %s -to %s %s" %(os.path.join(src_video,key + '.wav'),begin, end, os.path.join(tgt_video,key + "_" + str(cnt) + '.wav'))
            cnt += 1
            # os.system(ffmpeg_cmd)
            res = subprocess.call(ffmpeg_cmd, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                                  stderr=subprocess.PIPE)
        file_cnt += 1
        trans_progress(canvas_split, file_num, file_cnt)




'''
    将hh:MM:mm 转化为秒
'''
def time_trans(s_time):
    #print(s_time)
    return int(s_time.split(':')[0]) * 3600 + int(s_time.split(':')[1]) * 60  + int(s_time.split(':')[2])



if __name__ == "__main__":
    src_format = ''
    tgt_format = ''
    window = create_window('Audio trans', 500, 500)
    # 创建一个top frame
    top_frame = set_frame(window)
    # 创建四个子frame
    audio_trans_frame = set_frame(top_frame)
    tk.Label(audio_trans_frame,width=70,height=2,text="========================Audio trans========================").pack()
    first_frame = set_frame(top_frame)
    second_frame = set_frame(top_frame)
    thrid_frame = set_frame(top_frame)
    fourth_frame = set_frame(top_frame)

    # 给第一个frame添加元素,按照表格布局，1x4
    ## 创建标签，源文件路径
    label_src_file_path = tk.Label(first_frame, width=10, height=2, text="源文件路径")
    label_src_file_path.grid(row=0, column=0, padx=0, pady=0, ipadx=0, ipady=0)

    ## 创建单行输入框
    src_file_path = ''
    src_file_path = tk.StringVar(first_frame, value=src_file_path)
    entry_src_file_path_input = tk.Entry(first_frame, width=25, textvariable=src_file_path)
    entry_src_file_path_input.grid(row=0, column=1, padx=0, pady=0, ipadx=0, ipady=0)

    ## 创建文件类型选择框text=text, variable=file_type_var, value=value, command=command)
    radiobutton_path_type_file = tk.Radiobutton(first_frame, text="文件", value="file", command=select_src_file)
    radiobutton_path_type_file.grid(row=0, column=2, padx=0, pady=0, ipadx=0, ipady=0)
    radiobutton_path_type_folder = tk.Radiobutton(first_frame, text="文件夹", value="folder", command=select_src_folder)
    radiobutton_path_type_folder.grid(row=0, column=3, padx=0, pady=0, ipadx=0, ipady=0)

    ## 添加标签框---源格式
    label_src_audio_format = tk.Label(first_frame, width=10, height=2, text="源格式")
    label_src_audio_format.grid(row=0, column=4, padx=0, pady=0, ipadx=0, ipady=0)

    ## 添加listbox---源格式列表
    src_audio_format = tk.StringVar()     #创建变量，便于取值
    listbox_src_audio_options = ttk.Combobox(first_frame, textvariable=src_audio_format, width=3)
    listbox_src_audio_options['value'] = ("wav", "mp3")
    listbox_src_audio_options.current(0)
    listbox_src_audio_options.grid(row=0, column=5, padx=0, pady=0, ipadx=0, ipady=0)

    #给第二个frame添加控件
    ## 创建标签，源文件路径
    label_tgt_file_path = tk.Label(second_frame, width=10, height=2, text="目标文件路径")
    label_tgt_file_path.grid(row=0, column=0, padx=0, pady=0, ipadx=0, ipady=0)

    ## 创建单行输入框
    tgt_file_path = ''
    tgt_file_path = tk.StringVar(second_frame, value=tgt_file_path)
    entry_tgt_file_path_input = tk.Entry(second_frame, width=35, textvariable=tgt_file_path)
    entry_tgt_file_path_input.grid(row=0, column=1, padx=0, pady=0, ipadx=0, ipady=0)

    ## 创建选择button
    button_select = tk.Button(second_frame, text="Select", command=select_tgt_folder)
    button_select.grid(row=0, column=2, padx=0, pady=0, ipadx=0, ipady=0)

    ## 添加标签框---源格式
    label_tgt_audio_format = tk.Label(second_frame, width=10, height=2, text="目标格式")
    label_tgt_audio_format.grid(row=0, column=4, padx=0, pady=0, ipadx=0, ipady=0)

    ## 添加listbox---源格式列表
    tgt_audio_format = tk.StringVar()  # 创建变量，便于取值
    listbox_tgt_audio_options = ttk.Combobox(second_frame, textvariable=tgt_audio_format, width=3)
    listbox_tgt_audio_options['value'] = ("wav", "pcm")
    listbox_tgt_audio_options.current(0)
    listbox_tgt_audio_options.grid(row=0, column=5, padx=0, pady=0, ipadx=0, ipady=0)

    # 给thrid_format添加控件
    # 设置下载进度条
    tk.Label(thrid_frame, text='转换进度:', ).grid(row=0, column=0, padx=0, pady=10, ipadx=0, ipady=0)
    canvas = tk.Canvas(thrid_frame, width=430, height=22, bg="white")
    canvas.grid(row=0, column=1, padx=0, pady=0, ipadx=0, ipady=0)
    # tk.Button(thrid_frame, text='启动进度条', command=trans_progress).grid(row=0, column=3, padx=0, pady=0, ipadx=0, ipady=0)

    # 给fourth_format添加控件
    ## 创建开始button
    button_begin = tk.Button(fourth_frame, text="Begin", command=begin_trans)
    button_begin.grid(row=0, column=1, padx=20, pady=0, ipadx=0, ipady=0)

    ## 创建暂停button
    button_pause = tk.Button(fourth_frame, text="Pause", command=pause_trans)
    button_pause.grid(row=0, column=2, padx=20, pady=0, ipadx=0, ipady=0)

    ## 创建取消button
    button_cancel = tk.Button(fourth_frame, text="Cancel", command=cancel_trans)
    button_cancel.grid(row=0, column=3, padx=20, pady=0, ipadx=0, ipady=0)
#######################################################################################################################
    audio_split_frame = set_frame(top_frame)
    tk.Label(audio_split_frame, width=70, height=2, text="========================Audio split========================").pack()
    ## 创建四个子frame
    fifth_frame = set_frame(top_frame)
    sixth_frame = set_frame(top_frame)
    seventh_frame = set_frame(top_frame)
    eight_frame = set_frame(top_frame)

    ## 给fifth_frame添加元素,按照表格布局，1x4
    ### 创建标签，源文件路径
    label_src_file_path_split = tk.Label(fifth_frame, width=10, height=2, text="源文件路径")
    label_src_file_path_split.grid(row=0, column=0, padx=0, pady=0, ipadx=0, ipady=0)

    ### 创建单行输入框
    src_file_path_split = ''
    src_file_path_split = tk.StringVar(fifth_frame, value=src_file_path_split)
    entry_src_file_path_input_split = tk.Entry(fifth_frame, width=20, textvariable=src_file_path_split)
    entry_src_file_path_input_split.grid(row=0, column=1, padx=0, pady=0, ipadx=0, ipady=0)

    ### 创建文件类型选择框text=text, variable=file_type_var, value=value, command=command)
    radiobutton_path_type_file_split = tk.Radiobutton(fifth_frame, text="文件", value="file", command=select_src_file_split)
    radiobutton_path_type_file_split.grid(row=0, column=2, padx=0, pady=0, ipadx=0, ipady=0)
    radiobutton_path_type_folder_split = tk.Radiobutton(fifth_frame, text="文件夹", value="folder", command=select_src_folder_split)
    radiobutton_path_type_folder_split.grid(row=0, column=3, padx=0, pady=0, ipadx=0, ipady=0)

    ### 添加标签框---begin_time
    tk.Label(fifth_frame, width=10, height=2, text="begin").grid(row=0, column=4, padx=0, pady=0, ipadx=0, ipady=0)
    ### 添加listbox---begin time
    begin_time = ''
    begin_time = tk.StringVar(fifth_frame, value=begin_time)
    entry_begin_time = tk.Entry(fifth_frame, width=10, textvariable=begin_time)
    entry_begin_time.grid(row=0, column=5, padx=0, pady=0, ipadx=0, ipady=0)
    # 给sixth_frame添加控件
    ## 创建标签，源文件路径
    tk.Label(sixth_frame, width=10, height=2, text="目的文件路径").grid(row=0, column=0, padx=0, pady=0, ipadx=0, ipady=0)
    ## 创建单行输入框
    tgt_file_path_split = ''
    tgt_file_path_split = tk.StringVar(sixth_frame, value=tgt_file_path_split)
    entry_tgt_file_path_input_split = tk.Entry(sixth_frame, width=30, textvariable=tgt_file_path_split)
    entry_tgt_file_path_input_split.grid(row=0, column=1, padx=0, pady=0, ipadx=0, ipady=0)
    ## 创建选择button
    tk.Button(sixth_frame, text="Select", command=select_tgt_folder_split).grid(row=0, column=2, padx=0, pady=0, ipadx=0, ipady=0)
    ## 添加标签框---源格式
    tk.Label(sixth_frame, width=10, height=2, text="end:").grid(row=0, column=4, padx=0, pady=0, ipadx=0, ipady=0)
    ### 添加listbox---end time
    end_time = ''
    end_time = tk.StringVar(fifth_frame, value=end_time)
    entry_end_time = tk.Entry(sixth_frame, width=10, textvariable=end_time)
    entry_end_time.grid(row=0, column=5, padx=0, pady=0, ipadx=0, ipady=0)

    # 给seventh_format添加控件
    # 设置下载进度条
    tk.Label(seventh_frame, text='转换进度:', ).grid(row=0, column=0, padx=0, pady=10, ipadx=0, ipady=0)
    canvas_split = tk.Canvas(seventh_frame, width=430, height=22, bg="white")
    canvas_split.grid(row=0, column=1, padx=0, pady=0, ipadx=0, ipady=0)
    # tk.Button(thrid_frame, text='启动进度条', command=trans_progress).grid(row=0, column=3, padx=0, pady=0, ipadx=0, ipady=0)

    # 给eighth_frame添加控件
    ### 添加标签框---segment file
    tk.Label(eight_frame, width=10, height=2, text="分片文件").grid(row=0, column=0, padx=0, pady=0, ipadx=0, ipady=0)
    ### 添加listbox---segment file
    segment_file_path = ''
    segment_file_path = tk.StringVar(eight_frame, value=segment_file_path)
    entry_segment_file_path = tk.Entry(eight_frame, width=25, textvariable=segment_file_path)
    entry_segment_file_path.grid(row=0, column=1, padx=0, pady=0, ipadx=0, ipady=0)
    ### 添加segment file选择按钮
    tk.Button(eight_frame, text="Select", command=select_segment_file_split).grid(row=0, column=2, padx=0, pady=0,
                                                                                ipadx=0, ipady=0)
    ### 添加标签框---first audio name
    tk.Label(eight_frame, width=10, height=2, text="第一个音频名").grid(row=0, column=3, padx=2, pady=0, ipadx=0, ipady=0)
    ### 添加listbox---segment file
    first_audio_name = ''
    first_audio_name = tk.StringVar(eight_frame, value=first_audio_name)
    entry_first_audio_name = tk.Entry(eight_frame, width=10, textvariable=first_audio_name)
    entry_first_audio_name.grid(row=0, column=4, padx=0, pady=0, ipadx=0, ipady=0)
    ### 添加begin_split按钮
    tk.Button(eight_frame, text="Begin", command=begin_split).grid(row=0, column=5, padx=0, pady=0, ipadx=0, ipady=0)




    # 主窗口循环显示
    window.mainloop()

